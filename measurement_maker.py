#!/usr/bin/env python3
"""
Aggregate measurements.csv files into one Excel summary.

Usage:
    python aggregate_measurements.py /path/to/frame_out_full_vis \
        --output 20251120_실측_결과.xlsx
"""

import os
import argparse
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def parse_args():
    parser = argparse.ArgumentParser(
        description="Collect AVERAGE rows from measurements.csv in subdirectories "
                    "and create a single Excel summary file."
    )
    parser.add_argument(
        "root_dir",
        help="Root directory that contains per-schedule subdirectories "
             "with measurements.csv (e.g. frame_out_full_vis)."
    )
    parser.add_argument(
        "--output",
        "-o",
        default="measurement_summary.xlsx",
        help="Output Excel filename (default: measurement_summary.xlsx)."
    )
    return parser.parse_args()


def find_measurement_files(root_dir):
    """Return list of (schedule_id, csv_path) under root_dir."""
    results = []
    for name in sorted(os.listdir(root_dir)):
        subdir = os.path.join(root_dir, name)
        if not os.path.isdir(subdir):
            continue
        csv_path = os.path.join(subdir, "measurements.csv")
        if os.path.isfile(csv_path):
            results.append((name, csv_path))
    return results


def read_average_row(csv_path):
    """
    Read 'AVERAGE' row from measurements.csv and return
    (length_top_mm, length_bottom_mm, width_top_mm, width_bottom_mm) as float.
    """
    with open(csv_path, "r", newline="") as f:
        reader = csv.DictReader(f)
        avg_row = None
        for row in reader:
            # 'frame' 컬럼에 AVERAGE가 들어있는 행 찾기
            if str(row.get("frame", "")).strip().upper() == "AVERAGE":
                avg_row = row
                break

    if avg_row is None:
        raise ValueError(f"'AVERAGE' row not found in {csv_path}")

    def to_float(v):
        return float(str(v).strip())

    return (
        to_float(avg_row["P1-P2"]),
        to_float(avg_row["P3-P4"]),
        to_float(avg_row["P5-P6"]),
        to_float(avg_row["P7-P8"]),
    )


def create_excel(output_path, records):
    """
    records: list of dicts
      {
        "index": 1,
        "schedule_id": "202511230533537960",
        "p1p2": ...,
        "p3p4": ...,
        "p5p6": ...,
        "p7p8": ...,
      }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # --- Optional: some spacing rows (like screenshot) ---
    # Row 16: unit row (E16:H16 merged, "단위:mm")
    unit_row = 3
    ws.merge_cells(start_row=unit_row, start_column=6, end_row=unit_row, end_column=9)
    ws.cell(row=unit_row, column=6, value="단위:mm")

    # Row 17: header row
    header_row = 4
    headers = ["순번", "스케줄 ID", "P1-P2", "P3-P4", "P5-P6", "P7-P8", "비고"]
    # Put headers starting from column D (4)
    start_col = 4
    for i, h in enumerate(headers):
        ws.cell(row=header_row, column=start_col + i, value=h)

    # Data rows from row 18 downward
    row_idx = header_row + 1
    for rec in records:
        ws.cell(row=row_idx, column=start_col + 0, value=rec["index"])
        ws.cell(row=row_idx, column=start_col + 1, value=rec["schedule_id"])
        ws.cell(row=row_idx, column=start_col + 2, value=rec["p1p2"])
        ws.cell(row=row_idx, column=start_col + 3, value=rec["p3p4"])
        ws.cell(row=row_idx, column=start_col + 4, value=rec["p5p6"])
        ws.cell(row=row_idx, column=start_col + 5, value=rec["p7p8"])
        # "비고" column left blank
        row_idx += 1

    # Optional: set column widths a bit wider
    for col_offset, width in enumerate([6, 20, 14, 14, 14, 14, 18]):
        col_letter = get_column_letter(start_col + col_offset)
        ws.column_dimensions[col_letter].width = width

    wb.save(output_path)
    print(f"Saved Excel summary to: {output_path}")


def main():
    args = parse_args()
    root_dir = os.path.abspath(args.root_dir)

    files = find_measurement_files(root_dir)
    if not files:
        raise SystemExit(f"No measurements.csv found under: {root_dir}")

    records = []
    for idx, (schedule_id, csv_path) in enumerate(files, start=1):
        try:
            lt, lb, wt, wb = read_average_row(csv_path)
        except Exception as e:
            print(f"[WARN] Failed to read AVERAGE from {csv_path}: {e}")
            continue

        records.append(
            {
                "index": idx,
                "schedule_id": schedule_id,
                "p1p2": lt,
                "p3p4": lb,
                "p5p6": wt,
                "p7p8": wb,
            }
        )

    if not records:
        raise SystemExit("No valid AVERAGE rows were found.")

    create_excel(args.root_dir+args.output, records)


if __name__ == "__main__":
    main()
